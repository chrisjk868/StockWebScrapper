import openpyxl as xl
import yfinance as yf
import pandas as pd


def get_tickers(ws, col):
    process_tickers = dict()
    max_row = max((c.row for c in ws['B'] if c.value is not None))
    for row in range(2, max_row + 1):
        current_cell = "{}{}".format(col, row)
        process_tickers[ws[current_cell].value] = ['C' + current_cell[1:], 0]
    return process_tickers


def get_stock_prices(stocks):
    pd.set_option('display.max_columns', None)
    ticker_symbols = [*stocks]
    stocks_info = yf.Tickers(ticker_symbols)
    current_price = stocks_info.history(period='1d')['Close'][0:].to_numpy()[0]
    return current_price


def update_stock_prices(stock_index, close):
    i = 0
    for stock in stock_index.keys():
        stock_index[stock][1] = close[i]
        i += 1
    return stock_index


def write_to_excel(ws, stock_index):
    for stock in stock_index:
        ws[stock_index[stock][0]] = stock_index[stock][1]


def main():
    wb = xl.load_workbook('HK_2021.xlsx', data_only=True)
    worksheet_names = wb.sheetnames[1:]
    print(len(worksheet_names))
    for sheet_index in range(1, len(worksheet_names) + 1):
        wb.active = sheet_index
        print(wb.active.title)
        stock_dict = get_tickers(wb.active, 'B')
        closing_prices = get_stock_prices(stock_dict)
        updated_index = update_stock_prices(stock_dict, closing_prices)
        print(updated_index)
        write_to_excel(wb.active, updated_index)
        wb.save('HK_2021.xlsx')

    # Recovery Code:
    # ws = wb.active
    # stock_dict = get_tickers(ws, 'B')
    # closing_prices = get_stock_prices(stock_dict)
    # updated_index = update_stock_prices(stock_dict, closing_prices)
    # write_to_excel(ws, updated_index)
    # wb.save('HK_2021.xlsx')


if __name__ == '__main__':
    main()

